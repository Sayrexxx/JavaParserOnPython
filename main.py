import javalang
import re
import openpyxl
from openpyxl.utils import get_column_letter
import math


class JavaParser:
    def __init__(self):
        self.operators = {
            "Arithmetic Operators": r"(?<!\w)(?:\+\+|--|[+\-*/%])(?!\w)",
            "Assignment Operators": r"(?<!\w)(?:=|\+=|\-=|\*=|/=|%=|&=|\|=|\^=|<<=|>>=|>>>=)(?!\w)",
            "Comparison Operators": r"(?<!\w)(?:==|!=|>|<|>=|<=)(?!\w)",
            "Logic Operators": r"(?<!\w)(?:!|&&|\|\|)(?!\w)",
            "Bitwise Operators": r"(?<!\w)(?:&|\||\^|~|<<|>>|>>>)(?!\w)",
            "Ternary Operators": r"(?<!\w)\?(?!\w)",
            "Type Operators": r"(?<!\w)instanceof(?!\w)",
            "Control Flow Operators": r"(?<!\w)(?:if|else|switch|case|default|while|do|for|break|continue|return|try"
                                      r"|catch|finally|throw)(?!\w)",
            "Data Type Operators": r"(?<!\w)(?:boolean|byte|char|short|int|long|float|double)(?!\w)",
            "IO Operators": r"(?<!\w)(?:System\.out\.println|System\.out\.print|System\.err\.println|Scanner|System"
                            r"\.in)(?!\w)"
        }
        self.variables = {}

    def find_operators(self, code):
        operator_counts = {}
        for pattern in self.operators.values():
            operators = re.findall(pattern, code)
            for operator in operators:
                operator_counts[operator] = operator_counts.get(operator, 0) + 1

        return operator_counts

    @staticmethod
    def calculate_operator_count(code, operator_set):
        count = 0
        for operator in operator_set:
            count += code.count(operator)
        return count

    def count_variables(self, code):
        tree = javalang.parse.parse(code)
        for path, node in tree:
            if isinstance(node, javalang.tree.LocalVariableDeclaration):
                for declarator in node.declarators:
                    variable_name = declarator.name
                    if variable_name not in self.variables:
                        self.variables[variable_name] = 1
                    else:
                        self.variables[variable_name] += 1
            elif isinstance(node, javalang.tree.VariableDeclarator):
                variable_name = node.name
                if variable_name not in self.variables:
                    self.variables[variable_name] = 1
                else:
                    self.variables[variable_name] += 1
            elif isinstance(node, javalang.tree.MemberReference):
                variable_name = node.member
                if variable_name not in self.variables:
                    self.variables[variable_name] = 1
                else:
                    self.variables[variable_name] += 1

    def print_map(self, code):
        operator_counts = self.find_operators(code)
        self.count_variables(code)

        wb = openpyxl.Workbook()
        sheet = wb.active

        # Запись заголовков столбцов
        columns = ['Operator', 'count', 'Operand', 'count']
        for col_num, column_title in enumerate(columns, 1):
            col_letter = get_column_letter(col_num)
            sheet[f'{col_letter}1'] = column_title

        # Запись данных метрик Холстеда
        row_num1 = 2
        for operator, operator_count in operator_counts.items():
            sheet[f'A{row_num1}'] = operator
            sheet[f'B{row_num1}'] = operator_count
            row_num1 += 1

        row_num2 = 2
        for operand, operand_count in self.variables.items():
            sheet[f'C{row_num2}'] = operand
            sheet[f'D{row_num2}'] = operand_count
            row_num2 += 1
        # row_num = max(row_num1, row_num2)
        row_num = 30
        # Запись остальных метрик
        sheet[f'A{row_num}'] = 'n1'
        sheet[f'B{row_num}'] = 'N1'
        sheet[f'C{row_num}'] = 'n2'
        sheet[f'D{row_num}'] = 'N2'

        sheet[f'A{row_num + 1}'] = len(operator_counts.keys())
        sheet[f'B{row_num + 1}'] = sum(operator_counts.values())
        sheet[f'C{row_num + 1}'] = len(self.variables.keys())
        sheet[f'D{row_num + 1}'] = sum(self.variables.values())

        holsted_metrics1 = sheet[f'A{row_num + 1}'].value + sheet[f'C{row_num + 1}'].value
        holsted_metrics2 = sheet[f'B{row_num + 1}'].value + sheet[f'D{row_num + 1}'].value
        holsted_metrics3 = math.ceil(holsted_metrics2 * math.log2(holsted_metrics1))

        sheet[f'A{row_num + 3}'] = 'Словарь программы'
        sheet[f'B{row_num + 3}'] = holsted_metrics1
        sheet[f'A{row_num + 4}'] = 'Длина программы'
        sheet[f'B{row_num + 4}'] = holsted_metrics2
        sheet[f'A{row_num + 5}'] = 'Объем программы'
        sheet[f'B{row_num + 5}'] = holsted_metrics3

        # Сохранение документа Excel
        wb.save('holsted_metrics.xlsx')
        print("Результаты метрик Холстеда сохранены в файле holsted_metrics.xlsx")


parser = JavaParser()


with open("test.txt", "r") as f:
    java_code = f.read()

parser.print_map(java_code)